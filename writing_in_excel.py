import openpyxl

#File  workbook  sheets  rows  cells

#same data
# file = "C:/Users/anike/Downloads/selenium_excel.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet2"]
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value='Welcome'

# workbook.save(file)

#multiple data
file = "C:/Users/anike/Downloads/selenium_excel.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet3"]

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Aniket"
sheet.cell(1,3).value="QA"

sheet.cell(2,1).value=124
sheet.cell(2,2).value="Hima"
sheet.cell(2,3).value="Dev"

workbook.save(file)
