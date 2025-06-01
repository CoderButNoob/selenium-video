import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum,colNum).value

def writeData(file,sheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum,colNum).value=data

def fillGreenColor(file,sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    sheet.cell(row=rowNum, column=colNum).fill = greenFill
    workbook.save(file)

def fillRedColor(file, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.cell(row=rowNum, column=colNum).fill = redFill
    workbook.save(file)

