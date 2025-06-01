from openpyxl.styles.builtins import total

from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

#File  workbook  sheets  rows  cells
file = "C:/Users/anike/Downloads/selenium_excel.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

for row in range(1,rows+1):
    for col in range(1,cols+1):
        print(sheet.cell(row,col).value,end='       ')
    print()


